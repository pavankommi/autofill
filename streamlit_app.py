import streamlit as st # type: ignore
import openpyxl # type: ignore
from io import BytesIO

# Target columns to ensure exist
TARGET_COLUMNS = [
    "Vendor Code", "description", "ebay_description", "exclude_from_sitemap",
    "free_shipping", "installation_warranty_sku", "manage_stock",
    "manufacturer_status", "meta_description", "meta_keyword", "meta_title",
    "msrp_display_actual_price_type", "msrp_enabled", "options_container",
    "order_volume", "price", "product.attribute_set", "product.type",
    "product.websites", "short_description", "sku_prefix", "status",
    "stock.qty", "tax_class_id", "visibility", "condition", "supplier_lead_time"
]

# Default values per column
DEFAULT_VALUES = {
    "free_shipping": "1",
    "msrp_display_actual_price_type": "Use config",
    "tax_class_id": "Taxable Goods",
    "status": "Enabled",
    "condition": "New",
    "manufacturer_status": "Active Product",
    "order_volume": "Normal",
    "product.websites": "base",
    "exclude_from_sitemap": "No",
    "visibility": "Catalog, Search",
    "supplier_lead_time": "Ships Within 2-3 Business Days",
    "options_container": "Product Info Column",
    "stock.qty": "0",
    "product.type": "simple"
}

st.title("Excel Auto-Fill Based on 'sku' or 'mpn'")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    # Ensure 'sku' or 'mpn' exists
    if 'sku' not in headers and 'mpn' not in headers:
        st.error("Excel file must contain 'sku' or 'mpn' column.")
    else:
        # Add missing columns if needed
        for col in TARGET_COLUMNS:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers)).value = col

        col_map = {header: idx + 1 for idx, header in enumerate(headers)}
        sku_idx = col_map.get('sku', None)
        mpn_idx = col_map.get('mpn', None)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            has_sku = sku_idx and row[sku_idx - 1].value
            has_mpn = mpn_idx and row[mpn_idx - 1].value

            if has_sku or has_mpn:
                for col in TARGET_COLUMNS:
                    target_idx = col_map[col]
                    default_val = DEFAULT_VALUES.get(col, "")  # Empty string for unspecified columns
                    row[target_idx - 1].value = default_val

        st.success("Values updated. You can download the file.")

        output = BytesIO()
        wb.save(output)
        st.download_button(
            label="Download Updated Excel",
            data=output.getvalue(),
            file_name="updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

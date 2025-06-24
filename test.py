import streamlit as st # type: ignore
import pandas as pd # type: ignore
import openpyxl # type: ignore
from io import BytesIO

AUTO_FILL_COLUMNS = [
    "free_shipping", "msrp_enabled", "msrp_display_actual_price_type",
    "tax_class_id", "status", "condition", "manufacturer",
    "manufacturer_status", "order_volume", "product.websites",
    "exclude_from_sitemap", "manage_stock", "installation_warranty_sku",
    "visibility", "supplier_lead_time", "options_container",
    "product.attribute_set", "stock.qty", "product.type", "category.name"
]

st.title("Excel Auto-Fill (Keeps Original Formatting)")

uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])

if uploaded_file:
    wb = openpyxl.load_workbook(uploaded_file)
    ws = wb.active  # Assumes working on first sheet

    # Read header row
    headers = [cell.value for cell in ws[1]]

    if 'mpn' not in headers:
        st.error("Excel file must contain 'mpn' column.")
    else:
        mpn_idx = headers.index('mpn') + 1  # Excel columns start at 1

        # Ensure target columns exist, add them if missing
        for col in AUTO_FILL_COLUMNS:
            if col not in headers:
                headers.append(col)
                ws.cell(row=1, column=len(headers)).value = col

        # Create mapping for fast lookup
        col_map = {header: idx + 1 for idx, header in enumerate(headers)}

        # Process rows
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            mpn_val = row[mpn_idx - 1].value
            if mpn_val:
                for col in AUTO_FILL_COLUMNS:
                    target_idx = col_map[col]

                    # Skip first 2 columns entirely (assuming checkboxes)
                    if target_idx <= 2:
                        continue

                    row[target_idx - 1].value = "test"

        st.success("Values updated. You can download the file.")

        output = BytesIO()
        wb.save(output)
        st.download_button(
            label="Download Updated Excel",
            data=output.getvalue(),
            file_name="updated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
